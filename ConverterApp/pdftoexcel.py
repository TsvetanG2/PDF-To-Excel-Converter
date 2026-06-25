import os
import sys
import uuid
import json
import logging
import zipfile
from logging.handlers import RotatingFileHandler
from datetime import datetime
from functools import wraps
from concurrent.futures import ThreadPoolExecutor
import threading

import pdfplumber
import tabula
import fitz
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# Determine base directory (works for both script and frozen exe)
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Initialize Flask app
app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))

# Configuration - use absolute paths
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
ALLOWED_EXTENSIONS = {'pdf'}
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB limit
MAX_WORKERS = 4  # Thread pool size

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Setup logging
LOG_FOLDER = os.path.join(BASE_DIR, 'logs')
os.makedirs(LOG_FOLDER, exist_ok=True)

logger = logging.getLogger('pdf_converter')
logger.setLevel(logging.INFO)

file_handler = RotatingFileHandler(
    os.path.join(LOG_FOLDER, 'app.log'),
    maxBytes=10 * 1024 * 1024,  # 10MB
    backupCount=5
)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s - %(levelname)s - %(message)s'
))
logger.addHandler(file_handler)

console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter(
    '%(asctime)s - %(levelname)s - %(message)s'
))
logger.addHandler(console_handler)

# Thread pool for concurrent processing
executor = ThreadPoolExecutor(max_workers=MAX_WORKERS)

# Rate limiting storage (in production, use Redis)
rate_limit_storage = {}
rate_limit_lock = threading.Lock()
RATE_LIMIT_REQUESTS = 10  # requests per window
RATE_LIMIT_WINDOW = 60  # seconds

# Job storage for async processing
jobs = {}
jobs_lock = threading.Lock()

# Heartbeat tracking for auto-shutdown
last_heartbeat = datetime.now()
heartbeat_lock = threading.Lock()
HEARTBEAT_TIMEOUT = 30  # 30 seconds without heartbeat triggers shutdown
shutdown_event = threading.Event()


def get_client_ip():
    """Get client IP address."""
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    return request.remote_addr or '127.0.0.1'


def rate_limit(f):
    """Rate limiting decorator."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        client_ip = get_client_ip()
        current_time = datetime.now().timestamp()

        with rate_limit_lock:
            if client_ip not in rate_limit_storage:
                rate_limit_storage[client_ip] = []

            # Clean old requests
            rate_limit_storage[client_ip] = [
                t for t in rate_limit_storage[client_ip]
                if current_time - t < RATE_LIMIT_WINDOW
            ]

            if len(rate_limit_storage[client_ip]) >= RATE_LIMIT_REQUESTS:
                logger.warning(f"Rate limit exceeded for IP: {client_ip}")
                return jsonify({
                    'error': 'Rate limit exceeded. Please try again later.',
                    'retry_after': RATE_LIMIT_WINDOW
                }), 429

            rate_limit_storage[client_ip].append(current_time)

        return f(*args, **kwargs)
    return decorated_function


def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_tables_from_pdf(pdf_path):
    """Extract tables from all pages of a PDF file using tabula-py."""
    logger.info(f"Extracting tables from: {pdf_path}")
    all_tables = []
    try:
        with fitz.open(pdf_path) as doc:
            for page_num in range(len(doc)):
                tables = tabula.read_pdf(pdf_path, pages=page_num + 1, multiple_tables=True)
                if tables:
                    all_tables.extend(tables)
        logger.info(f"Extracted {len(all_tables)} tables from {pdf_path}")
    except Exception as e:
        logger.error(f"Error extracting tables from {pdf_path}: {str(e)}")
        raise
    return all_tables


def extract_pdf_content(pdf_path):
    """Extract text and tables from PDF."""
    logger.info(f"Extracting content from: {pdf_path}")
    text_content = ""
    table_data = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_content += page_text + "\n"

                tables = page.extract_tables()
                if tables:
                    table_data.extend(tables)
        logger.info(f"Extracted {len(text_content)} chars and {len(table_data)} tables")
    except Exception as e:
        logger.error(f"Error extracting content from {pdf_path}: {str(e)}")
        raise

    return text_content, table_data


def create_excel(text_content, table_data):
    """Create Excel workbook from text and table data."""
    logger.info("Creating Excel workbook")
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Content"

    # Precompute normalized table rows for O(1) exact match lookup
    table_match_index = {}
    for table_index, table in enumerate(table_data):
        for row in table:
            if row:
                normalized_row = " ".join(
                    "" if cell is None else str(cell).strip() for cell in row
                ).strip()
                if normalized_row:
                    table_match_index.setdefault(normalized_row, set()).add(table_index)

    ws['A1'] = ""
    text_lines = text_content.split('\n') if text_content else []
    current_row = 2
    max_row_position = 2
    table_data_used = set()

    for line in text_lines:
        stripped_line = line.strip()
        # O(1) exact match lookup instead of O(n*m) substring search
        matching_tables = table_match_index.get(stripped_line, set())
        line_in_table = bool(matching_tables)

        if not line_in_table:
            ws.cell(row=current_row, column=1, value=line)
            current_row += 1

        for table_index in matching_tables:
            if table_index not in table_data_used:
                table = table_data[table_index]
                for row_index, row_data in enumerate(table, start=current_row):
                    for col_index, value in enumerate(row_data, start=2):
                        ws.cell(row=row_index, column=col_index, value=value)
                max_row_position = max(max_row_position, current_row + len(table) + 1)
                table_data_used.add(table_index)
                current_row += len(table) + 1

    # Apply font to maintain formatting
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Times New Roman', size=11)

    logger.info("Excel workbook created successfully")
    return wb


def write_tables_to_excel(tables, excel_path):
    """Write tables data to an Excel file."""
    logger.info(f"Writing {len(tables)} tables to Excel: {excel_path}")
    workbook = Workbook()

    for table_num, table in enumerate(tables, start=1):
        sheet = workbook.create_sheet(title=f'Table_{table_num}')

        title_font = Font(name='Times New Roman', size=11, bold=True)
        info_font = Font(name='Times New Roman', size=10)
        alignment = Alignment(wrap_text=True, vertical='center')
        header_fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

        df = pd.DataFrame(table.values, columns=[col.title() if col else "" for col in table.columns])

        for col_num, column_header in enumerate(df.columns, start=1):
            if column_header:
                cell = sheet.cell(row=1, column=col_num, value=column_header)
                cell.font = title_font
                cell.alignment = alignment
                cell.fill = header_fill
                sheet.column_dimensions[get_column_letter(col_num)].width = max(len(str(column_header)) + 2, 10)

        for row_num, (_, row) in enumerate(df.iterrows(), start=2):
            for col_num, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.font = info_font
                cell.alignment = alignment
                current_width = sheet.column_dimensions[get_column_letter(col_num)].width
                sheet.column_dimensions[get_column_letter(col_num)].width = max(
                    current_width, len(str(value)) + 2
                )

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            max_text_length = max(len(str(cell.value)) if cell.value else 0 for cell in row)
            sheet.row_dimensions[row[0].row].height = 35 + (max_text_length // 50) * 5

    workbook.remove(workbook.active)
    workbook.save(excel_path)
    logger.info(f"Excel file saved: {excel_path}")


def create_csv(text_content, table_data, output_path):
    """Create CSV from extracted data."""
    logger.info(f"Creating CSV: {output_path}")
    all_data = []

    if text_content:
        for line in text_content.split('\n'):
            if line.strip():
                all_data.append([line.strip()])

    for table in table_data:
        for row in table:
            if row:
                all_data.append([str(cell) if cell else '' for cell in row])

    df = pd.DataFrame(all_data)
    df.to_csv(output_path, index=False, header=False)
    logger.info(f"CSV file saved: {output_path}")


def create_json(text_content, table_data, output_path):
    """Create JSON from extracted data."""
    logger.info(f"Creating JSON: {output_path}")
    result = {
        'text': text_content.split('\n') if text_content else [],
        'tables': []
    }

    for i, table in enumerate(table_data):
        table_dict = {
            'table_number': i + 1,
            'rows': []
        }
        for row in table:
            if row:
                table_dict['rows'].append([str(cell) if cell else '' for cell in row])
        result['tables'].append(table_dict)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    logger.info(f"JSON file saved: {output_path}")


def process_pdf_job(job_id, filepath, processing_option, output_format='xlsx'):
    """Process PDF in background thread."""
    logger.info(f"Starting job {job_id}: {processing_option}, format: {output_format}")

    try:
        with jobs_lock:
            jobs[job_id]['status'] = 'processing'
            jobs[job_id]['progress'] = 10

        base_name = os.path.splitext(os.path.basename(filepath))[0]

        if processing_option == 'allText':
            text_content, table_data = extract_pdf_content(filepath)

            with jobs_lock:
                jobs[job_id]['progress'] = 50

            if output_format == 'xlsx':
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{base_name}_output.xlsx')
                excel_file = create_excel(text_content, table_data)
                excel_file.save(output_path)
            elif output_format == 'csv':
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{base_name}_output.csv')
                create_csv(text_content, table_data, output_path)
            elif output_format == 'json':
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{base_name}_output.json')
                create_json(text_content, table_data, output_path)
            else:
                raise ValueError(f"Unsupported format: {output_format}")

        elif processing_option == 'tablesOnly':
            tables = extract_tables_from_pdf(filepath)

            with jobs_lock:
                jobs[job_id]['progress'] = 50

            if not tables:
                with jobs_lock:
                    jobs[job_id]['status'] = 'completed'
                    jobs[job_id]['progress'] = 100
                    jobs[job_id]['error'] = 'No tables found in the PDF.'
                delete_files(filepath)
                return

            if output_format == 'xlsx':
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{base_name}_tables.xlsx')
                write_tables_to_excel(tables, output_path)
            elif output_format == 'csv':
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{base_name}_tables.csv')
                all_data = []
                for table in tables:
                    df = pd.DataFrame(table.values, columns=table.columns)
                    all_data.append(df)
                if all_data:
                    combined = pd.concat(all_data, ignore_index=True)
                    combined.to_csv(output_path, index=False)
            elif output_format == 'json':
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{base_name}_tables.json')
                result = {'tables': []}
                for i, table in enumerate(tables):
                    df = pd.DataFrame(table.values, columns=table.columns)
                    result['tables'].append({
                        'table_number': i + 1,
                        'columns': list(table.columns),
                        'data': df.to_dict('records')
                    })
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(result, f, ensure_ascii=False, indent=2)
            else:
                raise ValueError(f"Unsupported format: {output_format}")
        else:
            raise ValueError(f"Invalid processing option: {processing_option}")

        with jobs_lock:
            jobs[job_id]['status'] = 'completed'
            jobs[job_id]['progress'] = 100
            jobs[job_id]['output_path'] = output_path
            jobs[job_id]['filename'] = os.path.basename(output_path)

        delete_files(filepath)
        logger.info(f"Job {job_id} completed successfully")

    except Exception as e:
        logger.error(f"Job {job_id} failed: {str(e)}")
        with jobs_lock:
            jobs[job_id]['status'] = 'failed'
            jobs[job_id]['error'] = str(e)
        delete_files(filepath)


def delete_files(*file_paths):
    """Delete files safely."""
    for path in file_paths:
        if path and os.path.exists(path):
            try:
                os.remove(path)
                logger.info(f"Deleted file: {path}")
            except PermissionError:
                logger.warning(f"Could not delete file (in use): {path}")
            except Exception as e:
                logger.error(f"Error deleting file {path}: {str(e)}")


@app.route('/')
def index():
    """Serve main page."""
    logger.info(f"Index page accessed from {get_client_ip()}")
    return render_template('pdftoexcel.html')


@app.route('/upload', methods=['POST'])
@rate_limit
def upload():
    """Handle file upload and start processing."""
    client_ip = get_client_ip()
    logger.info(f"Upload request from {client_ip}")

    if 'pdfFile' not in request.files:
        logger.warning(f"No file in request from {client_ip}")
        return jsonify({'error': 'No file provided'}), 400

    files = request.files.getlist('pdfFile')

    # Filter valid PDF files
    valid_files = [f for f in files if f.filename and allowed_file(f.filename)]

    if not valid_files:
        logger.warning(f"No valid PDF files from {client_ip}")
        return jsonify({'error': 'No valid PDF files provided'}), 400

    processing_option = request.form.get('processingOption', 'allText')
    output_format = request.form.get('outputFormat', 'xlsx')

    # Map output format to extension
    ext_map = {'xlsx': '.xlsx', 'csv': '.csv', 'json': '.json'}
    output_ext = ext_map.get(output_format, '.xlsx')

    try:
        output_files = []
        temp_pdf_files = []

        for file in valid_files:
            # Save PDF with unique prefix to avoid conflicts
            original_filename = secure_filename(file.filename)
            base_name = os.path.splitext(original_filename)[0]
            unique_id = str(uuid.uuid4())[:8]
            pdf_filename = f"{unique_id}_{original_filename}"
            pdf_filepath = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
            file.save(pdf_filepath)
            temp_pdf_files.append(pdf_filepath)
            logger.info(f"File saved: {pdf_filepath}")

            # Output filename uses original name
            output_filename = f"{base_name}{output_ext}"
            output_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{unique_id}_{output_filename}")

            if processing_option == 'allText':
                text_content, table_data = extract_pdf_content(pdf_filepath)

                if output_format == 'xlsx':
                    excel_file = create_excel(text_content, table_data)
                    excel_file.save(output_path)
                elif output_format == 'csv':
                    create_csv(text_content, table_data, output_path)
                elif output_format == 'json':
                    create_json(text_content, table_data, output_path)

            elif processing_option == 'tablesOnly':
                tables = extract_tables_from_pdf(pdf_filepath)
                if tables:
                    if output_format == 'xlsx':
                        write_tables_to_excel(tables, output_path)
                    elif output_format == 'csv':
                        all_data = []
                        for table in tables:
                            df = pd.DataFrame(table.values, columns=table.columns)
                            all_data.append(df)
                        if all_data:
                            combined = pd.concat(all_data, ignore_index=True)
                            combined.to_csv(output_path, index=False)
                    elif output_format == 'json':
                        result = {'tables': []}
                        for i, table in enumerate(tables):
                            df = pd.DataFrame(table.values, columns=table.columns)
                            result['tables'].append({
                                'table_number': i + 1,
                                'columns': list(table.columns),
                                'data': df.to_dict('records')
                            })
                        with open(output_path, 'w', encoding='utf-8') as f:
                            json.dump(result, f, ensure_ascii=False, indent=2)
                else:
                    logger.warning(f"No tables found in {original_filename}")
                    continue

            if os.path.exists(output_path):
                output_files.append((output_path, output_filename))

        # Clean up PDF files
        delete_files(*temp_pdf_files)

        if not output_files:
            return jsonify({'error': 'No files were converted successfully'}), 400

        # Single file - return directly with original name
        if len(output_files) == 1:
            output_path, output_filename = output_files[0]
            response = send_file(
                output_path,
                as_attachment=True,
                download_name=output_filename
            )

            @response.call_on_close
            def cleanup():
                delete_files(output_path)

            return response

        # Multiple files - create ZIP
        zip_filename = f"converted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        zip_path = os.path.join(app.config['UPLOAD_FOLDER'], zip_filename)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for output_path, output_filename in output_files:
                zipf.write(output_path, output_filename)

        # Clean up individual output files
        for output_path, _ in output_files:
            delete_files(output_path)

        logger.info(f"Created ZIP with {len(output_files)} files: {zip_path}")

        response = send_file(zip_path, as_attachment=True, download_name=zip_filename)

        @response.call_on_close
        def cleanup():
            delete_files(zip_path)

        return response

    except Exception as e:
        logger.error(f"Error processing upload from {client_ip}: {str(e)}")
        return jsonify({'error': f'Processing error: {str(e)}'}), 500


@app.route('/job/<job_id>')
def get_job_status(job_id):
    """Get status of async job."""
    with jobs_lock:
        if job_id not in jobs:
            return jsonify({'error': 'Job not found'}), 404

        job = jobs[job_id].copy()

    return jsonify(job)


@app.route('/job/<job_id>/download')
def download_job_result(job_id):
    """Download result of completed job."""
    with jobs_lock:
        if job_id not in jobs:
            return jsonify({'error': 'Job not found'}), 404

        job = jobs[job_id]

        if job['status'] != 'completed':
            return jsonify({'error': 'Job not completed'}), 400

        if 'error' in job:
            return jsonify({'error': job['error']}), 400

        output_path = job.get('output_path')

    if not output_path or not os.path.exists(output_path):
        return jsonify({'error': 'Output file not found'}), 404

    response = send_file(output_path, as_attachment=True)

    @response.call_on_close
    def cleanup():
        delete_files(output_path)
        with jobs_lock:
            if job_id in jobs:
                del jobs[job_id]

    return response


@app.route('/health')
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'active_jobs': len([j for j in jobs.values() if j['status'] == 'processing'])
    })


@app.route('/heartbeat', methods=['POST'])
def heartbeat():
    """Receive heartbeat from browser to keep server alive."""
    global last_heartbeat
    with heartbeat_lock:
        last_heartbeat = datetime.now()
    return jsonify({'status': 'ok'})


@app.route('/shutdown', methods=['POST'])
def shutdown():
    """Graceful shutdown endpoint (called when browser closes)."""
    logger.info("Shutdown requested via endpoint")
    shutdown_event.set()
    return jsonify({'status': 'shutting_down'})


def heartbeat_monitor():
    """Monitor heartbeat and shutdown server if no heartbeat received."""
    logger.info(f"Heartbeat monitor started (timeout: {HEARTBEAT_TIMEOUT}s)")

    while not shutdown_event.is_set():
        shutdown_event.wait(timeout=10)  # Check every 10 seconds

        with heartbeat_lock:
            elapsed = (datetime.now() - last_heartbeat).total_seconds()

        if elapsed > HEARTBEAT_TIMEOUT:
            logger.info(f"No heartbeat for {elapsed:.0f}s - initiating shutdown")
            shutdown_event.set()
            break

    logger.info("Heartbeat monitor: triggering server shutdown")
    os._exit(0)


def start_heartbeat_monitor():
    """Start the heartbeat monitor thread."""
    monitor_thread = threading.Thread(target=heartbeat_monitor, daemon=True)
    monitor_thread.start()
    return monitor_thread


@app.errorhandler(413)
def request_entity_too_large(error):
    """Handle file too large error."""
    logger.warning(f"File too large from {get_client_ip()}")
    return jsonify({
        'error': f'File too large. Maximum size is {MAX_CONTENT_LENGTH // (1024 * 1024)}MB'
    }), 413


@app.errorhandler(500)
def internal_error(error):
    import traceback
    traceback.print_exc()
    return jsonify({
        'error': str(error)
    }), 500

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    traceback.print_exc()

    return jsonify({
        'error': str(e)
    }), 500


if __name__ == "__main__":
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    logger.info(f"Starting PDF to Excel Converter (debug={debug})")
    app.run(host='0.0.0.0', port=5000, debug=debug, threaded=True)
